#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import re
import sys
import time
import urllib.parse
from pprint import pprint

import pandas as pd
import requests
from requests.exceptions import HTTPError
from tqdm import tqdm

ISISBN13 = re.compile("^978([0-9]{10})$")
ISJAN = re.compile("^19[12][0-9]{10}$")


def convert(csvfile=None):
    df = None
    orig = pd.read_csv(csvfile)

    isbn = [s.strip() for s in orig['ISBN'].tolist()]
    concat = orig['CONCAT'][0]

    isbn13 = [ib for ib in isbn if ISISBN13.match(ib)]
    jan = [ib for ib in isbn if ISJAN.match(ib)]

    others = set(isbn) - set(isbn13) - set(jan)

    remain_isbn13 = [s for s in list(others) if re.match("978([0-9]{9})", s)]
    isbn13 = isbn13 + [(s[:12] + '9') for s in remain_isbn13]

    others = others - set(remain_isbn13)

    print(f"{len(isbn13)} books.")
    print(f"others: {others}")

    m13 = re.compile("(978([0-9]{10}))")
    mjan = re.compile("(19[12]([0-9]{10}))")
    mj = re.compile("(49([0-9]{11}))")
    mi = re.compile("(45([0-9]{11}))")
    added_isbn13 = [s[0] for s in m13.findall(concat)]
    concat_remain = re.sub("(978([0-9]{10}))", "", concat)
    added_jan = [s[0] for s in mjan.findall(concat_remain)]
    concat_remain = re.sub("(19[12]([0-9]{10}))", "", concat_remain)
    added_mj = [s[0] for s in mj.findall(concat_remain)]
    concat_remain = re.sub("(49([0-9]{11}))", "", concat_remain)
    added_mi = [s[0] for s in mi.findall(concat_remain)]

    print(f"added isbn = {added_isbn13}")
    print(f"added jan = {added_jan}")
    print(f"added mj = {added_mj}")
    print(f"added mi = {added_mi}")

    print(f"remain = {concat_remain}")
    check = len(concat) - len(added_isbn13)*13 - len(added_jan) * \
        13 - len(added_mj)*13 - len(added_mi)*13
    print(f"check = {check}")

    # print(concat)

    df = pd.DataFrame({'ISBN': isbn13 + added_isbn13})

    return df


def js2row(di={}, json=None):
    item = json['items'][0]
    vinfo = item['volumeInfo']

    rowdict = {
        "image": "",  # dummy
        "title": vinfo['title'],
        "authors": ", ".join(vinfo['authors']),
        "categories":  ", ".join(vinfo['categories']),
        "description": vinfo['description'],
        "image url": vinfo['imageLinks']['smallThumbnail'],
        "isbn13": "".join([v['identifier'] for v in vinfo['industryIdentifiers'] if v['type'] == 'ISBN_13']),
    }

    for k, v in di.items():
        dik = di.get(k, [])
        dik.append(rowdict[k])
        di[k] = dik

    return di


def fetch(isbn=None):
    try:
        url = f"isbn:{isbn}"
        enc_url = "https://www.googleapis.com/books/v1/volumes?q=" + \
            urllib.parse.quote(url)
        response = requests.get(enc_url)
        response.raise_for_status()
        # access JSOn content
        return response.json()

    except HTTPError as http_err:
        print(f'HTTP error occurred: {http_err}')
    except Exception as err:
        print(f'Other error occurred: {err}')


def to_excel(df=None, excelfile=None):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils.cell import get_column_letter
    columns = ["image", "title", "authors", "categories",
               "description", "image url", "isbn13"]

    wb = Workbook()
    ws = wb.active

    ws[1] = columns
    for i in range(len(columns)):
        ws[f"{get_column_letter(i + 1)}1"] = columns[i]

    for idx, url in enumerate(df['image url']):
        ws.add_image(Image(url), f"{get_column_letter(c + 1)}{idx + 2}")
        time.sleep(1.2)
        for j, k in enumerate(columns[1:]):
            ws[f"{get_column_letter(j + 2)}{idx + 2}"] = df[k][idx]

    wb.save(excelfile)


def main(csvfile=os.path.join(os.path.dirname(__file__), "..", "csv", "books.csv"),
         excelfile=os.path.join(os.path.dirname(__file__), "..", "csv", "books.xlsx")):
    df = convert(csvfile=csvfile)
    df.to_csv(os.path.join(os.path.dirname(__file__), "..", "csv", "isbn.csv"))
    print(df)

    data = {}
    for idx, i in enumerate(tqdm(list(df['ISBN']))):
        data = js2row(di=data, json=fetch(isbn=df['ISBN'][0]))
        time.sleep(1.2)

    df_out = pd.DataFrame(data, columns=[
                          "image", "title", "authors", "categories", "description", "image url", "isbn13"])

    to_excel(df=df_out, excelfile=excelfile)


if __name__ == "__main__":
    main()
